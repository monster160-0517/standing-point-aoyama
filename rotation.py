import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import random
import re
import json
from datetime import date
from html import escape
from io import BytesIO
from urllib.parse import quote
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 1. 페이지 설정
st.set_page_config(page_title="GENTLEMONSTER AOYAMA", layout="wide")

st.markdown("""
    <style>
    .meal-bg { background-color: #ffff00 !important; color: black !important; font-weight: bold; }
    </style>
    """, unsafe_allow_html=True)

st.title("GENTLEMONSTER AOYAMA")
st.caption("フロアスケジュール / Floor schedule")

# 🔗 AOYAMA 전용 시트 설정
STORE_NAME = "GENTLEMONSTER AOYAMA"
SHEET_ID = "1XcXSvokpLlkWnTQtqs-Zlz6lLISXJ9Zy14kEOIOUGlA"
DB_SHEET_GID = "0"
TO_SHEET_GID = "410487706"
RAW_TO_DISPLAY = {
    "식사": "休憩",
    "2回目休憩": "2回休",
    "도슨트": "ETC",
    "1층 유동": "1F-OP",
    "2층 유동": "2F-OP",
}
COLUMN_ALIASES = {
    "name": ["이름", "氏名", "名前", "Name"],
    "type": ["구분", "区分", "雇用区分", "Type"],
    "start_time": ["출근시간", "出勤時間", "Start Time"],
    "end_time": ["퇴근시간", "退勤時間", "End Time"],
    "lunch": ["점심", "昼休憩", "Lunch"],
    "dinner": ["저녁", "夕方休憩", "Dinner"],
    "meal_time": ["식사시간", "食事時間", "食事時間1H", "Meal Time"],
    "second_break": ["2回目休憩", "２回目休憩", "第二休憩", "Second Break"],
    "counter_flag": ["카운터여부", "カウンター可否", "レジ可能有無", "Counter Eligible"],
    "flex_flag": ["유동여부", "流動可否", "OP可能有無", "Float Eligible", "OP"],
}
DOCENT_COLUMN_TOKENS = ["도슨트", "ドーセント", "DOCENT", "Docent", "ETC"]
TIME_SLOT_STEP_MINUTES = 30
MEAL_DURATION_MINUTES = 60
SECOND_BREAK_DURATION_MINUTES = 30

st.sidebar.markdown(f"**🏠 {STORE_NAME}**")

def load_sheet_data(sheet_id, gid=None, sheet_name=None, show_errors=True):
    if sheet_name:
        encoded_sheet_name = quote(sheet_name)
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={encoded_sheet_name}"
    else:
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        df = pd.read_csv(url, skip_blank_lines=True, dtype=str, keep_default_na=False)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.fillna("").replace(r'\.0$', '', regex=True)
        return df
    except Exception as e:
        if show_errors:
            st.error(f"シートの読み込みに失敗しました / Failed to load sheet: {e}")
        return pd.DataFrame()

def load_first_available_sheet(sheet_id, sheet_names):
    for sheet_name in sheet_names:
        df = load_sheet_data(sheet_id, sheet_name=sheet_name, show_errors=False)
        if not df.empty:
            return df
    return pd.DataFrame()

def find_column(columns, aliases, fallback=None):
    for alias in aliases:
        if alias in columns:
            return alias
    return fallback if fallback is not None else aliases[0]

def find_docent_columns(columns):
    return [
        column for column in columns
        if any(token.lower() in str(column).lower() for token in DOCENT_COLUMN_TOKENS)
    ]

def count_staff_rows(data):
    if data.empty:
        return 0
    name_col = find_column(data.columns, COLUMN_ALIASES["name"], None)
    type_col = find_column(data.columns, COLUMN_ALIASES["type"], None)
    if not name_col or not type_col or name_col not in data.columns or type_col not in data.columns:
        return 0

    count = 0
    for _, row in data.iterrows():
        name = str(row.get(name_col, "")).strip()
        staff_type = str(row.get(type_col, "")).strip().lower()
        if name and any(token in staff_type for token in ['정직', '파트', '正社員', 'アルバイト', 'ft', 'pt', 'full', 'part']):
            count += 1
    return count

db_df = load_sheet_data(SHEET_ID, DB_SHEET_GID)
to_df = load_sheet_data(SHEET_ID, TO_SHEET_GID)
docent_df = load_first_available_sheet(SHEET_ID, ["도슨트", "ドーセント", "Docent"])

if db_df.empty: st.stop()

def get_clean_time(val):
    val = str(val).strip()
    if not val: return None
    nums = re.findall(r'\d+', val)
    if not nums:
        return None

    hour = int(nums[0])
    minute = int(nums[1]) if len(nums) > 1 else 0
    lower_val = val.lower()
    is_pm = "오후" in val or "pm" in lower_val
    is_am = "오전" in val or "am" in lower_val

    if is_pm and hour < 12:
        hour += 12
    elif is_am and hour == 12:
        hour = 0

    return f"{hour:02d}:{minute:02d}"

def get_minutes_from_time(val, default=None):
    clean = get_clean_time(val)
    if not clean:
        return default
    hour, minute = clean.split(":")
    return int(hour) * 60 + int(minute)

def minutes_to_time(total_minutes):
    hour = total_minutes // 60
    minute = total_minutes % 60
    return f"{hour:02d}:{minute:02d}"

def align_to_step(total_minutes, step_minutes=TIME_SLOT_STEP_MINUTES):
    remainder = total_minutes % step_minutes
    if remainder == 0:
        return total_minutes
    return total_minutes + (step_minutes - remainder)

def build_slot_range(start_minutes, end_minutes, step_minutes=TIME_SLOT_STEP_MINUTES):
    if start_minutes is None or end_minutes is None or end_minutes <= start_minutes:
        return []
    return [minutes_to_time(minute) for minute in range(start_minutes, end_minutes, step_minutes)]

def build_duration_slots(start_time, duration_minutes):
    start_minutes = get_minutes_from_time(start_time)
    if start_minutes is None:
        return []
    end_minutes = start_minutes + duration_minutes
    return build_slot_range(start_minutes, end_minutes)

def build_expanded_to_rows(source_df):
    expanded_rows = {}
    if source_df.empty:
        return expanded_rows
    for _, row in source_df.iterrows():
        base_minutes = get_minutes_from_time(row.iloc[0])
        if base_minutes is None:
            continue
        expanded_rows[minutes_to_time(base_minutes)] = row
        expanded_rows[minutes_to_time(base_minutes + TIME_SLOT_STEP_MINUTES)] = row
    return expanded_rows

def build_work_range(in_val, out_val, default_in=None, default_out=None):
    in_minutes = get_minutes_from_time(in_val, default_in)
    out_minutes = get_minutes_from_time(out_val, default_out)
    if in_minutes is None or out_minutes is None:
        return None, None, None
    if out_minutes <= in_minutes:
        return None, in_minutes, out_minutes
    return set(build_slot_range(in_minutes, out_minutes)), in_minutes, out_minutes

def parse_time_list(value):
    text = str(value).strip()
    if not text:
        return []
    chunks = re.split(r'[,/\n;]+', text)
    times = []
    for chunk in chunks:
        clean = get_clean_time(chunk)
        if clean:
            times.append(clean)
    return sorted(set(times))

def translate_zone_name(zone_name):
    text = str(zone_name).strip()
    if not text:
        return text

    upper = text.upper()
    if upper == "COUNTER":
        return "カウンター"
    if upper == "1F":
        return "1Fゾーン"
    if upper == "2F":
        return "2Fゾーン"
    if upper == "OP":
        return "OP"
    if upper.startswith("ETC"):
        return text

    translated = text
    translated = translated.replace("카운터", "カウンター")
    translated = translated.replace("유동", "OP")
    translated = translated.replace("PHOTO", "フォト")
    return translated

def to_display_value(value):
    text = normalize_schedule_value(value)
    if text in RAW_TO_DISPLAY:
        return RAW_TO_DISPLAY[text]
    return translate_zone_name(text)

def is_counter_zone(zone_name):
    zone = str(zone_name).upper()
    return "카운터" in zone or "カウンター" in zone or "COUNTER" in zone or "1F-C" in zone or "2F-C" in zone

def is_flexible_zone(zone_name):
    zone = str(zone_name)
    return "유동" in zone or "流動" in zone or "FLOAT" in zone.upper() or str(zone_name).strip().upper() == "OP"

def get_zone_category(zone_name):
    zone = str(zone_name).upper()
    if is_counter_zone(zone_name):
        return "counter"
    if is_flexible_zone(zone_name):
        return "flex"
    if "2F" in zone or "2층" in zone:
        return "2f"
    if "1F" in zone or "1층" in zone:
        return "1f"
    return "other"

def get_floor_bucket(zone_name):
    if not zone_name:
        return None
    zone = str(zone_name)
    upper = zone.upper()
    if "1층" in zone or "1F" in upper:
        return "1f"
    if "2층" in zone or "2F" in upper:
        return "2f"
    if "카운터" in zone or "COUNTER" in upper:
        if "カウンター" in zone:
            if "2" in zone:
                return "2f"
            if "1" in zone:
                return "1f"
            return "counter"
        # default to whichever floor indicator is included
        if "2" in zone:
            return "2f"
        if "1" in zone:
            return "1f"
        return "counter"
    if "유동" in zone or "流動" in zone:
        if "2" in zone:
            return "2f"
        if "1" in zone:
            return "1f"
    return None

def get_zone_priority(zone_name):
    if is_counter_zone(zone_name):
        return 0
    if is_flexible_zone(zone_name):
        return 2
    return 1

def is_w_zone(zone_name):
    zone = str(zone_name).upper()
    return bool(re.search(r'(^|[^A-Z])W($|[^A-Z])', zone))

def is_photo_zone(zone_name):
    zone = str(zone_name).upper()
    return "PHOTO" in zone or "フォト" in str(zone_name)

def is_b1_zone(zone_name):
    zone = str(zone_name).upper()
    return "B1" in zone

def is_docent_zone(zone_name):
    zone = str(zone_name).strip()
    return "도슨트" in zone or "ドーセント" in zone or "DOCENT" in zone.upper() or zone.upper().startswith("ETC")

def get_special_zone_group(zone_name):
    if is_w_zone(zone_name):
        return "w"
    return None

def is_docent_assignment(value):
    text = str(value).strip()
    return text in {"도슨트", "ドーセント", "ETC"} or text.upper().startswith("ETC")

def normalize_schedule_value(value):
    text = str(value)
    if text.strip() == "":
        return ""
    return text

def is_enabled_flag(value):
    return any(x in str(value).lower() for x in ['o', 'y', '1', 'v', '예', '可', 'yes', 'true'])

def pick_best_staff(zone_name, pool, previous_assignments):
    if not pool:
        return None

    candidates = list(pool)
    random.shuffle(candidates)

    if is_flexible_zone(zone_name):
        non_flexible_last = [
            name for name in candidates
            if not is_flexible_zone(previous_assignments.get(name))
        ]
        if non_flexible_last:
            candidates = non_flexible_last
    else:
        from_flexible_last = [
            name for name in candidates
            if is_flexible_zone(previous_assignments.get(name))
        ]
        if from_flexible_last:
            candidates = from_flexible_last

    non_consecutive = [
        name for name in candidates if previous_assignments.get(name) != zone_name
    ]
    if non_consecutive:
        candidates = non_consecutive

    return candidates[0] if candidates else None

# --- 기본 데이터 파싱 ---
def get_initial_staff(data):
    type_col = find_column(data.columns, COLUMN_ALIASES["type"], "구분")
    name_col = find_column(data.columns, COLUMN_ALIASES["name"], "이름")
    start_col = find_column(data.columns, COLUMN_ALIASES["start_time"], "출근시간")
    end_col = find_column(data.columns, COLUMN_ALIASES["end_time"], "퇴근시간")
    lunch_col = find_column(data.columns, COLUMN_ALIASES["lunch"], "점심")
    dinner_col = find_column(data.columns, COLUMN_ALIASES["dinner"], "저녁")
    meal_col = find_column(data.columns, COLUMN_ALIASES["meal_time"], "식사시간")
    second_break_col = find_column(data.columns, COLUMN_ALIASES["second_break"], "2回目休憩")
    counter_col = find_column(data.columns, COLUMN_ALIASES["counter_flag"], "카운터여부")
    flex_col = find_column(data.columns, COLUMN_ALIASES["flex_flag"], "유동여부")
    docent_cols = find_docent_columns(data.columns)
    res = []
    for _, row in data.iterrows():
        name = str(row.get(name_col, "")).strip()
        stype = str(row.get(type_col, "")).strip()
        if name and any(kw in stype.lower() for kw in ['정직', '파트', '正社員', 'アルバイト', 'part', 'full', 'ft', 'pt']):
            docent_times = []
            for docent_col in docent_cols:
                docent_times.extend(parse_time_list(row.get(docent_col, "")))
            res.append({
                "original_name": name,
                "type": '정직' if any(kw in stype.lower() for kw in ['정직', '正社員', 'full', 'ft']) else '파트',
                "in": get_clean_time(row.get(start_col, '11')),
                "out": get_clean_time(row.get(end_col, '21')),
                "meal1": get_clean_time(row.get(lunch_col, '')),
                "meal2": get_clean_time(row.get(dinner_col, '')),
                "meal_p": get_clean_time(row.get(meal_col, '')),
                "second_break_override": get_clean_time(row.get(second_break_col, '')),
                "can_counter": is_enabled_flag(row.get(counter_col, 'X')),
                "can_flexible": is_enabled_flag(row.get(flex_col, 'X')),
                "docent_times": sorted(set(docent_times)),
            })
    return res

def get_docent_schedule(staff_rows, extra_data):
    primary_schedule = {
        staff["original_name"]: list(staff.get("docent_times", []))
        for staff in staff_rows
        if staff.get("docent_times")
    }

    if primary_schedule:
        return primary_schedule

    schedule = {}

    if extra_data.empty:
        return schedule

    name_col = find_column(extra_data.columns, COLUMN_ALIASES["name"], "이름")
    docent_cols = find_docent_columns(extra_data.columns)
    if not docent_cols:
        return schedule

    for _, row in extra_data.iterrows():
        name = str(row.get(name_col, "")).strip()
        docent_times = []
        for docent_col in docent_cols:
            docent_times.extend(parse_time_list(row.get(docent_col, "")))
        if name:
            merged = sorted(set(schedule.get(name, []) + docent_times))
            if merged:
                schedule[name] = merged

    return schedule

raw_staff = get_initial_staff(db_df)
docent_schedule = get_docent_schedule(raw_staff, docent_df)

# --- 사이드바: 파트타이머 상세 조정 ---
st.sidebar.header("🕹️ 人員管理 / Staffing")
use_docent_schedule = st.sidebar.checkbox("🎤 ETC予定を反映 / Apply ETC schedule", value=True)
with st.sidebar.expander("🎤 ETCタブ / ETC tab", expanded=False):
    st.caption("`ETC1`, `ETC2`, `ETC3` 列を自動認識します / Automatically detects ETC1/2/3 columns.")
    if docent_schedule:
        for docent_name, docent_times in sorted(docent_schedule.items()):
            st.write(f"{docent_name}: {', '.join(docent_times)}")
    else:
        st.caption("ETC時間の記録がありません / No ETC times found.")

pt_list = [s for s in raw_staff if s['type'] == '파트']

pt_input_defaults = {
    s["original_name"]: {
        "in": s["in"] or "",
        "out": s["out"] or "",
        "meal": s["meal_p"] or "",
    }
    for s in pt_list
}
pt_input_signature = json.dumps(
    {
        "store": STORE_NAME,
        "defaults": pt_input_defaults,
    },
    ensure_ascii=False,
    sort_keys=True,
)

if st.session_state.get("pt_input_signature") != pt_input_signature:
    for pt_name, defaults in pt_input_defaults.items():
        st.session_state[f"in_{pt_name}"] = defaults["in"]
        st.session_state[f"out_{pt_name}"] = defaults["out"]
        st.session_state[f"meal_{pt_name}"] = defaults["meal"]
    st.session_state["pt_input_signature"] = pt_input_signature

default_pt_names = [s['original_name'] for s in pt_list if s.get('in') and s.get('out')]
selected_pt_names = st.sidebar.multiselect(
    "⏱️ 出勤アルバイト選択 / Select part-timers on duty",
    [s['original_name'] for s in pt_list],
    default=default_pt_names,
)

final_staff_configs = []

# 1. 정직원 처리 (조 이름 포함)
for s in [x for x in raw_staff if x['type'] == '정직']:
    work_range, in_minutes, out_minutes = build_work_range(s['in'], s['out'])
    if work_range is None:
        continue
    tag = "(A Shift)" if in_minutes <= 10 * 60 else "(B Shift)"
    s['display_name'] = f"{s['original_name']}{tag}"
    if s.get('meal_p'):
        s['meals'] = build_duration_slots(s['meal_p'], MEAL_DURATION_MINUTES)
    else:
        meal_slots = []
        for meal_time in [s['meal1'], s['meal2']]:
            meal_slots.extend(build_duration_slots(meal_time, MEAL_DURATION_MINUTES))
        s['meals'] = sorted(set(meal_slots), key=lambda value: get_minutes_from_time(value, 0))
    s['first_meal_time'] = s.get('meal_p') or s.get('meal1') or s.get('meal2')
    s['docent_times'] = docent_schedule.get(s['original_name'], []) if use_docent_schedule else []
    s['work_range'] = work_range
    s['work_start_minutes'] = in_minutes
    s['work_end_minutes'] = out_minutes
    s['second_break_override'] = s.get('second_break_override')
    s['second_breaks'] = []
    s['in'] = minutes_to_time(in_minutes)
    s['out'] = minutes_to_time(out_minutes)
    final_staff_configs.append(s)

# 2. 파트타이머 처리 (조 이름 제외 + 사이드바 조정값 반영)
if selected_pt_names:
    st.sidebar.markdown("---")
    st.sidebar.subheader("📋 アルバイト時間調整 / Part-timer hours")
    for pt_name in selected_pt_names:
        pt_origin = next(s for s in pt_list if s['original_name'] == pt_name)
        with st.sidebar.expander(f"👤 {pt_name}"):
            c1, c2 = st.columns(2)
            new_in = c1.text_input("出勤 / In", key=f"in_{pt_name}")
            new_out = c2.text_input("退勤 / Out", key=f"out_{pt_name}")
            new_meal = st.text_input("食事 / Meal", key=f"meal_{pt_name}")

            pt_copy = pt_origin.copy()
            work_range, in_minutes, out_minutes = build_work_range(new_in, new_out)

            if work_range is None:
                st.warning(f"{pt_name}: 出勤/退勤時間を再確認してください / Please check the in/out times.")
                continue

            pt_copy['display_name'] = pt_name # 조 태그 없음
            pt_copy['in'] = minutes_to_time(in_minutes)
            pt_copy['out'] = minutes_to_time(out_minutes)
            pt_copy['meal_p'] = get_clean_time(new_meal)
            pt_copy['meals'] = build_duration_slots(pt_copy['meal_p'], MEAL_DURATION_MINUTES)
            pt_copy['first_meal_time'] = pt_copy['meal_p']
            pt_copy['docent_times'] = docent_schedule.get(pt_copy['original_name'], []) if use_docent_schedule else []
            pt_copy['work_range'] = work_range
            pt_copy['work_start_minutes'] = in_minutes
            pt_copy['work_end_minutes'] = out_minutes
            pt_copy['second_break_override'] = pt_copy.get('second_break_override')
            pt_copy['second_breaks'] = []
            final_staff_configs.append(pt_copy)

def assign_second_breaks(staff_configs):
    meal_groups = {}
    for staff in staff_configs:
        staff['second_breaks'] = []
        first_meal_time = staff.get('first_meal_time')
        if first_meal_time:
            meal_groups.setdefault(first_meal_time, []).append(staff)

    if not meal_groups:
        return

    last_first_meal_end = max(
        get_minutes_from_time(meal_time, 0) + MEAL_DURATION_MINUTES
        for meal_time in meal_groups
    )
    cursor_minutes = align_to_step(last_first_meal_end)
    occupied_break_slots = set()

    def find_break_slot(staff, requested_start_minutes):
        start_minutes = max(requested_start_minutes, last_first_meal_end)
        candidate_minutes = align_to_step(start_minutes)
        while candidate_minutes + SECOND_BREAK_DURATION_MINUTES <= staff.get('work_end_minutes', 0):
            candidate_slot = minutes_to_time(candidate_minutes)
            if (
                candidate_slot in staff.get('work_range', set())
                and candidate_slot not in staff.get('meals', [])
                and candidate_slot not in staff.get('docent_times', [])
                and candidate_slot not in occupied_break_slots
            ):
                return candidate_slot
            candidate_minutes += TIME_SLOT_STEP_MINUTES
        return None

    for meal_time in sorted(meal_groups, key=lambda value: get_minutes_from_time(value, 0)):
        for staff in meal_groups[meal_time]:
            override_slot = staff.get('second_break_override')
            second_break_slot = None
            if override_slot:
                override_minutes = get_minutes_from_time(override_slot)
                if override_minutes is not None:
                    second_break_slot = find_break_slot(staff, override_minutes)
            if second_break_slot is None:
                second_break_slot = find_break_slot(staff, cursor_minutes)
            if second_break_slot:
                staff['second_breaks'] = [second_break_slot]
                occupied_break_slots.add(second_break_slot)
                cursor_minutes = get_minutes_from_time(second_break_slot, cursor_minutes) + SECOND_BREAK_DURATION_MINUTES

assign_second_breaks(final_staff_configs)

config_signature = json.dumps(
    {
        "store": STORE_NAME,
        "staff": [
            {
                "name": s["display_name"],
                "type": s["type"],
                "in": s.get("in"),
                "out": s.get("out"),
                "meals": s.get("meals", []),
                "second_breaks": s.get("second_breaks", []),
                "docent_times": s.get("docent_times", []),
                "can_counter": s.get("can_counter", False),
                "can_flexible": s.get("can_flexible", False),
            }
            for s in final_staff_configs
        ],
        "use_docent_schedule": use_docent_schedule,
        "docent_schedule": docent_schedule,
    },
    ensure_ascii=False,
    sort_keys=True,
)

if st.session_state.get("config_signature") != config_signature:
    st.session_state.pop("result_df", None)
    st.session_state["config_signature"] = config_signature

staff_config_by_name = {s["display_name"]: s for s in final_staff_configs}

def enforce_priority_slots(df):
    enforced = df.copy()
    for staff_name, staff_config in staff_config_by_name.items():
        if staff_name not in enforced.index:
            continue
        for docent_slot in staff_config.get("docent_times", []):
            if docent_slot in enforced.columns:
                enforced.at[staff_name, docent_slot] = "도슨트"
        for meal_slot in staff_config.get("meals", []):
            if meal_slot in enforced.columns:
                enforced.at[staff_name, meal_slot] = "식사"
        for break_slot in staff_config.get("second_breaks", []):
            if break_slot in enforced.columns:
                enforced.at[staff_name, break_slot] = "2回目休憩"
    return enforced

# --- 로테이션 엔진 ---
def run_rotation():
    working_names = [s['display_name'] for s in final_staff_configs]
    staff_start_times = [s.get('work_start_minutes') for s in final_staff_configs if s.get('work_start_minutes') is not None]
    staff_end_times = [s.get('work_end_minutes') for s in final_staff_configs if s.get('work_end_minutes') is not None]
    to_time_values = [
        get_minutes_from_time(value)
        for value in to_df[to_df.columns[0]].tolist()
        if get_minutes_from_time(value) is not None
    ]
    min_schedule_minutes = min(staff_start_times + to_time_values) if (staff_start_times or to_time_values) else 11 * 60
    max_schedule_minutes = max(staff_end_times + [value + 60 for value in to_time_values]) if (staff_end_times or to_time_values) else 21 * 60
    all_time_slots = build_slot_range(min_schedule_minutes, max_schedule_minutes)
    schedule_df = pd.DataFrame(index=all_time_slots, columns=working_names).fillna("-")
    all_zones = [c for c in to_df.columns if c != to_df.columns[0]]

    # TO 시트에 실제 존재하는 유동 구역만 사용
    flex_zones_in_to = [z for z in all_zones if is_flexible_zone(z)]
    flex_1f = next((z for z in flex_zones_in_to if "1" in z), None)
    flex_2f = next((z for z in flex_zones_in_to if "2" in z), None)

    staff_lookup = {s['display_name']: s for s in final_staff_configs}
    previous_assignments = {n: None for n in working_names}
    floor_state = {n: {"floor": None, "count": 0} for n in working_names}
    floor_1f_total = {n: 0 for n in working_names}  # 1층 총 배정 횟수
    special_zone_history = {n: {} for n in working_names}
    w_zone_hours = {n: 0 for n in working_names}
    counter_assignment_total = {n: 0 for n in working_names}
    counter_zone_assignment_total = {n: {} for n in working_names}
    counter_consecutive_hours = {n: 0 for n in working_names}
    assignment_locks = {n: None for n in working_names}
    MAX_1F = 6
    MAX_W_HOURS = 4
    MAX_CONSECUTIVE_COUNTER_HOURS = 2
    MAX_SAME_FLOOR_SLOTS = 4

    expanded_to_rows = build_expanded_to_rows(to_df)
    counter_eligible_slot_capacity = {}
    for name, staff in staff_lookup.items():
        unavailable_slots = set(staff.get("meals", [])) | set(staff.get("second_breaks", [])) | set(staff.get("docent_times", []))
        eligible_slots = [slot for slot in staff.get("work_range", set()) if slot not in unavailable_slots]
        counter_eligible_slot_capacity[name] = max(len(eligible_slots), 1)

    def get_zone_identity(zone_name):
        return str(zone_name).strip().upper()

    def is_part_timer(name):
        staff = staff_lookup.get(name, {})
        return staff.get("type") == "파트"

    def can_assign_zone(name, zone_name):
        special_group = get_special_zone_group(zone_name)
        zone_identity = get_zone_identity(zone_name)

        if previous_assignments.get(name) == zone_name:
            return False

        if special_group:
            previous_zone = special_zone_history[name].get(special_group)
            if previous_zone and previous_zone != zone_identity:
                return False

        if is_w_zone(zone_name) and w_zone_hours[name] >= MAX_W_HOURS:
            return False

        if is_counter_zone(zone_name) and counter_consecutive_hours[name] >= MAX_CONSECUTIVE_COUNTER_HOURS:
            return False

        return True

    def record_zone_assignment(name, zone_name):
        special_group = get_special_zone_group(zone_name)
        if special_group:
            special_zone_history[name][special_group] = get_zone_identity(zone_name)
        if is_w_zone(zone_name):
            w_zone_hours[name] += 1
        if is_counter_zone(zone_name):
            counter_assignment_total[name] += 1
            zone_identity = get_zone_identity(zone_name)
            counter_zone_assignment_total[name][zone_identity] = (
                counter_zone_assignment_total[name].get(zone_identity, 0) + 1
            )
            counter_consecutive_hours[name] += 1
        else:
            counter_consecutive_hours[name] = 0

    def pick_counter_staff(zone_name, candidates):
        if not candidates:
            return None

        zone_identity = get_zone_identity(zone_name)
        shuffled_candidates = list(candidates)
        random.shuffle(shuffled_candidates)

        def counter_score(name):
            total_capacity = counter_eligible_slot_capacity.get(name, 1)
            counter_load_ratio = counter_assignment_total[name] / total_capacity
            return (
                round(counter_load_ratio, 4),
                counter_assignment_total[name],
                counter_zone_assignment_total[name].get(zone_identity, 0),
                counter_consecutive_hours[name],
                floor_1f_total[name],
            )

        best_score = min(counter_score(name) for name in shuffled_candidates)
        best_candidates = [
            name for name in shuffled_candidates
            if counter_score(name) == best_score
        ]

        return best_candidates[0] if best_candidates else None

    def update_floor_state(name, zone):
        floor = get_floor_bucket(zone)
        state = floor_state[name]
        if not floor:
            state["floor"] = None
            state["count"] = 0
            return
        if floor == "1f":
            floor_1f_total[name] += 1
        if state["floor"] == floor:
            state["count"] += 1
        else:
            state["floor"] = floor
            state["count"] = 1

    def can_assign_same_floor(name, floor):
        if not floor:
            return True
        state = floor_state[name]
        return not (state["floor"] == floor and state["count"] >= MAX_SAME_FLOOR_SLOTS)

    def parse_zone_capacity(raw_value):
        raw = str(raw_value).strip()
        if raw in ["", "0", "-", "nan"]:
            return 0
        return int(raw.split('-')[0]) if '-' in raw else int(float(raw or 0))

    def should_lock_to_next_half_hour(slot):
        slot_minutes = get_minutes_from_time(slot)
        if slot_minutes is None:
            return False
        return slot_minutes % 60 == 0

    def build_zone_assignment_plan(to_row):
        counter_pass = []
        other_first_pass = []
        other_extra_pass = []

        for zone_name in all_zones:
            capacity = parse_zone_capacity(to_row[zone_name])
            if capacity <= 0 or is_docent_zone(zone_name):
                continue
            if is_counter_zone(zone_name):
                counter_pass.extend([zone_name] * capacity)
            else:
                other_first_pass.append(zone_name)
                if capacity > 1:
                    other_extra_pass.extend([zone_name] * (capacity - 1))

        return counter_pass + other_first_pass + other_extra_pass

    def choose_staff_for_zone(zone_name, pool_names, current_zone_count, ignore_1f_limit=False):
        is_first_coverage_assignment = current_zone_count == 0
        zone_is_1f = get_floor_bucket(zone_name) == "1f"
        eligible = [
            n for n in pool_names
            if not (is_counter_zone(zone_name) and not staff_lookup[n]["can_counter"])
            and not (is_flexible_zone(zone_name) and not staff_lookup[n]["can_flexible"])
            and not (zone_is_1f and floor_1f_total[n] >= MAX_1F and not ignore_1f_limit)
            and can_assign_zone(n, zone_name)
        ]
        if is_b1_zone(zone_name):
            part_timer_eligible = [n for n in eligible if is_part_timer(n)]
            if part_timer_eligible:
                eligible = part_timer_eligible
        zone_floor = get_floor_bucket(zone_name)
        floor_filtered = [n for n in eligible if can_assign_same_floor(n, zone_floor)]
        working_candidates = floor_filtered or eligible
        if is_counter_zone(zone_name):
            chosen = pick_counter_staff(zone_name, working_candidates)
        else:
            chosen = pick_best_staff(
                zone_name,
                working_candidates,
                previous_assignments,
            )
        if not chosen and is_photo_zone(zone_name) and is_first_coverage_assignment:
            photo_fallback = [
                n for n in pool_names
                if can_assign_zone(n, zone_name)
            ]
            chosen = pick_best_staff(
                zone_name,
                photo_fallback,
                previous_assignments,
            )
        return chosen

    def assign_staff_to_zone(slot, zone_name, chosen_name, zone_assigned_count, create_followup_lock=True):
        schedule_df.at[slot, chosen_name] = zone_name
        previous_assignments[chosen_name] = zone_name
        record_zone_assignment(chosen_name, zone_name)
        zone_assigned_count[zone_name] = zone_assigned_count.get(zone_name, 0) + 1
        pool.remove(chosen_name)
        update_floor_state(chosen_name, zone_name)
        if create_followup_lock and should_lock_to_next_half_hour(slot):
            assignment_locks[chosen_name] = {"zone": zone_name, "remaining_slots": 1}
        else:
            assignment_locks[chosen_name] = None

    for slot in all_time_slots:
        pool = []
        for s in final_staff_configs:
            if slot in s["meals"]:
                schedule_df.at[slot, s['display_name']] = "식사"
                assignment_locks[s['display_name']] = None
                previous_assignments[s['display_name']] = None
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
            elif slot in s.get("second_breaks", []):
                schedule_df.at[slot, s['display_name']] = "2回目休憩"
                assignment_locks[s['display_name']] = None
                previous_assignments[s['display_name']] = None
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
            elif slot in s.get("docent_times", []):
                schedule_df.at[slot, s['display_name']] = "도슨트"
                assignment_locks[s['display_name']] = None
                previous_assignments[s['display_name']] = None
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
            elif slot in s["work_range"]:
                pool.append(s['display_name'])
            else:
                schedule_df.at[slot, s['display_name']] = " "
                assignment_locks[s['display_name']] = None
                previous_assignments[s['display_name']] = None
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
        
        random.shuffle(pool)
        to_row = expanded_to_rows.get(slot)
        
        if to_row is not None:
            zone_assignment_plan = build_zone_assignment_plan(to_row)
            zone_assigned_count = {}
            zone_required_capacity = {
                z: parse_zone_capacity(to_row[z])
                for z in all_zones
                if parse_zone_capacity(to_row[z]) > 0 and not is_docent_zone(z)
            }

            locked_names = [name for name in list(pool) if assignment_locks.get(name)]
            for locked_name in locked_names:
                lock_info = assignment_locks.get(locked_name)
                if not lock_info:
                    continue
                locked_zone = lock_info["zone"]
                assign_staff_to_zone(
                    slot,
                    locked_zone,
                    locked_name,
                    zone_assigned_count,
                    create_followup_lock=False,
                )

            for z in zone_assignment_plan:
                current_count = zone_assigned_count.get(z, 0)
                chosen = choose_staff_for_zone(z, pool, current_count)
                if not chosen:
                    continue
                assign_staff_to_zone(slot, z, chosen, zone_assigned_count)

            # First pass can miss valid assignments because earlier choices tighten constraints.
            # Try one more sweep with the remaining pool before sending people to flexible / "-".
            if pool:
                remaining_zones = []
                for zone_name in all_zones:
                    required = zone_required_capacity.get(zone_name, 0)
                    assigned = zone_assigned_count.get(zone_name, 0)
                    if required > assigned:
                        remaining_zones.extend([zone_name] * (required - assigned))

                for z in remaining_zones:
                    chosen = choose_staff_for_zone(z, pool, zone_assigned_count.get(z, 0))
                    if not chosen:
                        chosen = choose_staff_for_zone(
                            z,
                            pool,
                            zone_assigned_count.get(z, 0),
                            ignore_1f_limit=True,
                        )
                    if not chosen:
                        continue
                    assign_staff_to_zone(slot, z, chosen, zone_assigned_count)

            flexible_pool = [n for n in pool if staff_lookup[n]["can_flexible"]]
            inflexible_pool = [n for n in pool if not staff_lookup[n]["can_flexible"]]

            for n in flexible_pool: # TO 시트에 있는 유동 구역만 사용
                if not flex_1f and not flex_2f:
                    # 유동 구역 자체가 없으면 미배정
                    schedule_df.at[slot, n] = "-"
                    previous_assignments[n] = None
                    floor_state[n]["floor"] = None
                    floor_state[n]["count"] = 0
                    continue

                current_assignments = [
                    val for val in schedule_df.loc[slot].tolist()
                    if str(val).strip() not in ["-", "", " ", "식사"]
                ]
                f1_cnt = sum(1 for val in current_assignments if get_zone_category(val) == "1f")
                f2_cnt = sum(1 for val in current_assignments if get_zone_category(val) == "2f")

                can_1f = flex_1f and floor_1f_total[n] < MAX_1F
                if can_1f and flex_2f:
                    flexible_zone = flex_1f if f1_cnt <= f2_cnt else flex_2f
                    if previous_assignments.get(n) == flexible_zone:
                        flexible_zone = flex_2f if flexible_zone == flex_1f else flex_1f
                    # 1층 선택됐는데 한도 초과면 2층으로
                    if flexible_zone == flex_1f and floor_1f_total[n] >= MAX_1F:
                        flexible_zone = flex_2f
                elif flex_2f:
                    flexible_zone = flex_2f
                elif can_1f:
                    flexible_zone = flex_1f
                else:
                    flexible_zone = flex_2f  # 한도 초과 시 2층으로 강제

                schedule_df.at[slot, n] = flexible_zone
                previous_assignments[n] = flexible_zone
                record_zone_assignment(n, flexible_zone)
                update_floor_state(n, flexible_zone)

            for n in inflexible_pool:
                schedule_df.at[slot, n] = "-"
                previous_assignments[n] = None
                floor_state[n]["floor"] = None
                floor_state[n]["count"] = 0
                counter_consecutive_hours[n] = 0
    return schedule_df

if st.sidebar.button("🚀 ローテーション自動生成 / Generate rotation", width="stretch"):
    st.session_state.result_df = run_rotation()

# --- 화면 출력 ---
if 'result_df' in st.session_state:
    res = st.session_state.result_df
    st.write(f"### 📅 [{STORE_NAME}] フロアローテーション / Floor rotation")
    st.caption("下の表で修正すると、下部の共有ボードにすぐ反映されます / Edits below are reflected immediately in the shared board preview.")
    raw_display_df = res.transpose().map(normalize_schedule_value)
    raw_display_df.index.name = "직원명"
    display_to_raw = {}
    for raw_value in sorted(set(str(val).strip() for val in raw_display_df.values.flatten()) | set(zone_columns := [c for c in to_df.columns if c != to_df.columns[0]]) | set(RAW_TO_DISPLAY.keys())):
        display_value = to_display_value(raw_value)
        if display_value:
            display_to_raw[display_value] = raw_value
    display_df = raw_display_df.map(to_display_value)
    display_df.index.name = "氏名"
    editor_df = display_df.reset_index()
    editor_df = editor_df[["氏名"] + [c for c in editor_df.columns if c != "氏名"]]
    zone_choices = set(display_to_raw.keys())
    zone_choices.update(str(val).strip() for val in display_df.values.flatten() if str(val).strip())
    zone_choices.update(to_display_value(val) for val in ["식사", "2回目休憩", "도슨트", "1층 유동", "2층 유동", "-", ""])
    zone_choices = sorted(choice for choice in zone_choices if choice != "")
    column_settings = {
        col: (
            st.column_config.SelectboxColumn(options=zone_choices, width="small")
            if col != "氏名"
            else st.column_config.TextColumn(label="氏名", disabled=True, width="medium")
        )
        for col in editor_df.columns
    }
    edited_editor_df = st.data_editor(
        editor_df,
        width="stretch",
        height=450,
        column_config=column_settings,
        hide_index=True,
        num_rows="fixed",
        key="rotation_editor",
    )
    edited_display_df = edited_editor_df.copy()
    edited_display_df["氏名"] = edited_display_df["氏名"].astype(str).str.strip()
    edited_display_df = edited_display_df.set_index("氏名")
    edited_display_df.index.name = "氏名"
    edited_display_df = edited_display_df.reindex(columns=display_df.columns)
    edited_display_df = edited_display_df.map(normalize_schedule_value)
    edited_raw_df = edited_display_df.map(lambda value: display_to_raw.get(str(value).strip(), normalize_schedule_value(value)))
    edited_raw_df.index = raw_display_df.index
    edited_raw_df.index.name = "직원명"
    edited_raw_df = enforce_priority_slots(edited_raw_df)
    edited_display_df = edited_raw_df.map(to_display_value)
    edited_display_df.index.name = "氏名"
    csv_bytes = edited_display_df.to_csv(index=True).encode('utf-8-sig')
    file_name = f"rotation_{STORE_NAME}_{date.today():%Y%m%d}"

    def parse_required_count(raw_value):
        raw = str(raw_value).strip()
        if raw in ["", "0", "-", "nan"]:
            return 0
        return int(raw.split('-')[0]) if '-' in raw else int(float(raw or 0))

    def build_zone_coverage_summary(df):
        coverage_rows = []
        total_empty_zones = 0
        total_remaining_to = 0
        affected_times = set()
        expanded_to_rows = build_expanded_to_rows(to_df)

        active_zones = []
        for zone in zone_columns:
            if is_docent_zone(zone):
                continue
            for slot in df.columns:
                to_row = expanded_to_rows.get(slot)
                if to_row is None:
                    continue
                required = parse_required_count(to_row[zone])
                if required > 0:
                    active_zones.append(zone)
                    break
        active_zones = list(dict.fromkeys(active_zones))

        for zone in active_zones:
            row = {"zone": zone}
            for slot in df.columns:
                to_row = expanded_to_rows.get(slot)
                if to_row is None:
                    row[slot] = None
                    continue

                required = parse_required_count(to_row[zone])
                if required <= 0:
                    row[slot] = None
                    continue

                assigned = sum(1 for value in df[slot].tolist() if str(value).strip() == zone)
                remaining_to = max(required - assigned, 0)

                if assigned == 0:
                    total_empty_zones += 1
                    affected_times.add(slot)
                if remaining_to > 0:
                    total_remaining_to += remaining_to

                row[slot] = {
                    "assigned": assigned,
                    "to": required,
                    "remaining_to": remaining_to,
                    "is_empty": assigned == 0,
                }

            coverage_rows.append(row)

        return coverage_rows, total_empty_zones, total_remaining_to, len(affected_times)

    def get_staff_color(name):
        s_info = next((s for s in final_staff_configs if s['display_name'] == name), None)
        if not s_info:
            return "#111827"
        if "(A Shift)" in name:
            return "#f97316"
        if "(B Shift)" in name:
            return "#2563eb"
        if s_info["type"] == '정직':
            return "#1d4ed8"
        return "#059669"

    def get_zone_background(value):
        text = str(value)
        low = text.lower()
        if "카운터" in low or "counter" in low:
            return "#ede9fe"
        if "2층" in low or "2f" in low:
            return "#fee2e2"
        if "1층" in low or "1f" in low:
            return "#dbeafe"
        return ""

    def excel_color(hex_color):
        return hex_color.replace("#", "").upper()

    def style_rotation_worksheet(ws, df):
        thin_side = Side(style="thin", color="DDDDDD")
        header_fill = PatternFill(fill_type="solid", fgColor=excel_color("#f8f9fa"))
        meal_fill = PatternFill(fill_type="solid", fgColor=excel_color("#fff5ba"))
        second_break_fill = PatternFill(fill_type="solid", fgColor=excel_color("#dcfce7"))
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center_alignment
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_idx, staff in enumerate(df.index, start=2):
            name_cell = ws.cell(row=row_idx, column=1)
            name_cell.font = Font(bold=True, color=excel_color(get_staff_color(staff)))
            name_cell.alignment = center_alignment
            name_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            for col_idx, value in enumerate(df.loc[staff], start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                if str(value) in {"식사", "休憩 / Meal"}:
                    cell.fill = meal_fill
                    continue
                if str(value) in {"2回目休憩", "2回目休憩 / 2nd Break"}:
                    cell.fill = second_break_fill
                    continue
                if is_docent_assignment(value):
                    cell.fill = PatternFill(fill_type="solid", fgColor=excel_color("#fde68a"))
                    continue

                zone_color = get_zone_background(value)
                if zone_color:
                    cell.fill = PatternFill(fill_type="solid", fgColor=excel_color(zone_color))

        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 18

        for col_idx in range(2, len(df.columns) + 2):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = 12

    with BytesIO() as buf:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            edited_display_df.to_excel(writer, index=True, sheet_name="rotation")
            style_rotation_worksheet(writer.book["rotation"], edited_display_df)
        buf.seek(0)
        excel_bytes = buf.getvalue()

    def build_table(df):
        table_html = "<div class='table-scroll'><table class='rotation-table'>"
        table_html += "<thead><tr><th>氏名</th>"
        for time in df.columns:
            table_html += f"<th>{escape(str(time))}</th>"
        table_html += "</tr></thead><tbody>"
        for staff, row in df.iterrows():
            color = get_staff_color(staff)
            table_html += f"<tr><td class='staff-name' style='color: {color};'>{escape(str(staff))}</td>"
            for _, val in row.items():
                text = normalize_schedule_value(val)
                bg = ""
                if text in {"식사", "休憩 / Meal"}:
                    bg = "background-color: #fff5ba;"
                elif text in {"2回目休憩", "2回目休憩 / 2nd Break"}:
                    bg = "background-color: #dcfce7;"
                elif is_docent_assignment(text):
                    bg = "background-color: #fde68a;"
                else:
                    zone_color = get_zone_background(text)
                    if zone_color:
                        bg = f"background-color: {zone_color};"
                table_html += f"<td style='{bg}'>{escape(text)}</td>"
            table_html += "</tr>"
        table_html += "</tbody></table></div>"
        return table_html

    def build_zone_coverage_table(coverage_rows, time_slots):
        table_html = "<div class='table-scroll coverage-scroll'><table class='rotation-table coverage-table'>"
        table_html += "<thead><tr><th>ゾーン</th>"
        for slot in time_slots:
            table_html += f"<th>{escape(str(slot))}</th>"
        table_html += "</tr></thead><tbody>"

        for row in coverage_rows:
            table_html += f"<tr><td class='staff-name zone-name'>{escape(translate_zone_name(row['zone']))}</td>"
            for slot in time_slots:
                cell = row.get(slot)
                if cell is None:
                    table_html += "<td class='coverage-cell inactive'>-</td>"
                    continue

                classes = ["coverage-cell"]
                if cell["is_empty"]:
                    classes.append("empty")
                elif cell["remaining_to"] > 0:
                    classes.append("partial")
                else:
                    classes.append("filled")

                table_html += (
                    f"<td class='{' '.join(classes)}'>"
                    f"<div class='coverage-assigned'>{cell['assigned']}名</div>"
                    f"<div class='coverage-required'>TO {cell['to']}名</div>"
                    "</td>"
                )
            table_html += "</tr>"

        table_html += "</tbody></table></div>"
        return table_html

    table_styles = (
        "<style>"
        ".table-scroll{overflow:auto;background:#fff;border:1px solid #ddd;border-radius:12px;}"
        ".rotation-table{width:max-content;min-width:100%;border-collapse:collapse;font-size:0.82rem;line-height:1.15;}"
        ".rotation-table th,.rotation-table td{border:1px solid #ddd;padding:5px 6px;text-align:center;vertical-align:middle;white-space:nowrap;min-width:58px;}"
        ".rotation-table thead th{position:sticky;top:0;background:#f8f9fa;z-index:3;}"
        ".rotation-table .staff-name{position:sticky;left:0;background:#fff;font-weight:700;z-index:2;min-width:122px;max-width:122px;}"
        ".rotation-table thead th:first-child{left:0;z-index:4;}"
        ".coverage-scroll{margin:12px 0 20px;}"
        ".coverage-table .zone-name{font-weight:700;color:#111827;}"
        ".coverage-cell{min-width:72px;background:#ffffff;}"
        ".coverage-cell.inactive{background:#f8fafc;color:#94a3b8;}"
        ".coverage-cell.filled{background:#f0fdf4;}"
        ".coverage-cell.partial{border:2px solid #eab308 !important;background:#fef9c3;}"
        ".coverage-cell.empty{border:2px solid #dc2626 !important;background:#fff1f2;}"
        ".coverage-assigned{font-size:0.84rem;font-weight:700;color:#111827;}"
        ".coverage-required{margin-top:3px;font-size:0.68rem;color:#64748b;}"
        "</style>"
    )
    coverage_rows, total_empty_zones, total_remaining_to, affected_times = build_zone_coverage_summary(edited_raw_df)
    table_html = build_table(edited_display_df)
    coverage_table_html = build_zone_coverage_table(coverage_rows, edited_display_df.columns)
    page_html = "<!doctype html><html lang='ja'><head><meta charset='utf-8'/><title>モバイル共有ボード / Mobile share board</title>"
    page_html += (
        "<style>"
        "html,body{height:100%;margin:0;padding:0;background:#f8fafc;font-family:'Pretendard','Noto Sans KR',sans-serif;overflow:hidden;}"
        ".page-wrap{display:flex;flex-direction:column;height:100vh;padding:8px;box-sizing:border-box;gap:6px;overflow:hidden;}"
        "h1{margin:0;font-size:0.95rem;line-height:1.1;flex:0 0 auto;}"
        ".fit-stage{flex:1 1 auto;overflow:hidden;border-radius:10px;}"
        ".fit-frame{display:block;transform-origin:top left;}"
        ".table-scroll{overflow:visible;border-radius:10px;box-shadow:none;}"
        ".rotation-table{font-size:0.7rem;table-layout:fixed;background:#fff;}"
        ".rotation-table th,.rotation-table td{padding:3px 4px;min-width:46px;}"
        ".rotation-table .staff-name{min-width:104px;max-width:104px;}"
        ".rotation-table thead th{position:static;}"
        ".rotation-table .staff-name{position:static;}"
        "</style>"
        f"{table_styles}"
    )
    page_html += (
        "</head><body><div class='page-wrap'>"
        "<h1>モバイル共有ボード / Mobile share board</h1>"
        "<div class='fit-stage'><div class='fit-frame' id='fit-frame'>"
    )
    page_html += table_html
    page_html += (
        "</div></div></div>"
        "<script>"
        "function fitRotationTable(){"
        "const stage=document.querySelector('.fit-stage');"
        "const frame=document.getElementById('fit-frame');"
        "const table=document.querySelector('.rotation-table');"
        "if(!stage||!frame||!table)return;"
        "frame.style.width='auto';"
        "frame.style.height='auto';"
        "frame.style.margin='0 auto';"
        "frame.style.transform='scale(1)';"
        "const stageWidth=Math.max(stage.clientWidth-4,1);"
        "const stageHeight=Math.max(stage.clientHeight-4,1);"
        "const tableWidth=table.offsetWidth;"
        "const tableHeight=table.offsetHeight;"
        "if(!tableWidth||!tableHeight)return;"
        "const scale=Math.min(stageWidth/tableWidth, 1);"
        "frame.style.width=(tableWidth*scale)+'px';"
        "frame.style.height=(tableHeight*scale)+'px';"
        "frame.style.transform='scale('+scale+')';"
        "}"
        "window.addEventListener('load', fitRotationTable);"
        "window.addEventListener('resize', fitRotationTable);"
        "</script>"
        "</body></html>"
    )
    safe_page_html = page_html.replace("</script>", "<\\/script>")

    widget_html = f"""
    <div style='margin-bottom:8px;'>
        <button style='border:0; padding:10px 16px; font-weight:600; background:#111827; color:#fff; border-radius:8px; cursor:pointer;'
                onclick='openLargeRotation()'>🖥️ 拡大表示 / Open large view</button>
    </div>
    <script>
    const largeRotationContent = {json.dumps(safe_page_html)};
    function openLargeRotation() {{
        const win = window.open('', '_blank');
        if (!win) return;
        win.document.write(largeRotationContent);
        win.document.close();
    }}
    </script>
    """
    st.markdown(table_styles, unsafe_allow_html=True)
    st.markdown("### 🚨 ゾーン別配置チェック / Zone coverage check")
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("空きゾーン数 / Empty zones", total_empty_zones)
    metric_col2.metric("残りTO合計 / Remaining TO", total_remaining_to)
    metric_col3.metric("影響時間帯 / Affected hours", affected_times)
    st.caption("TOは最大配置可能人数です。`0名`は赤、TO未充足は黄色で強調します / TO means maximum capacity. Empty cells are red and underfilled cells are yellow.")
    st.markdown(coverage_table_html, unsafe_allow_html=True)
    st.write("---")
    st.markdown("### 🎨 カラー現況表 / Color schedule")
    st.caption("上の編集表の内容がすぐ反映される読み取り専用プレビューです / Read-only preview synced with the editable table above.")
    st.markdown(table_html, unsafe_allow_html=True)
    st.write("---")
    st.markdown("### 📸 モバイル共有ボード / Mobile share board")
    components.html(widget_html, height=110)
    st.write("---")
    st.markdown("### 📥 ダウンロード / Downloads")
    st.download_button("📥 現在の配置をダウンロード (CSV) / Download current schedule", data=csv_bytes, file_name=f"{file_name}.csv", mime="text/csv")
    st.download_button(
        "📥 現在の配置をダウンロード (Excel) / Download current schedule",
        data=excel_bytes,
        file_name=f"{file_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
